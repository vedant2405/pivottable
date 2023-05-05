import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# select the worksheet with the pivot table
ws = wb['PivotTable']

# select the pivot table
pivot = ws['A1'].pivotTable

# set the border style
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

# set the font style
font = Font(name='Calibri', size=12)

# set the alignment
align = Alignment(horizontal='center')

# set the fill color
fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# apply styles to the pivot table cells
for row in pivot.rows:
    for cell in row:
        cell.border = border
        cell.font = font
        cell.alignment = align
        cell.fill = fill

# save the workbook
wb.save('example_formatted.xlsx')
